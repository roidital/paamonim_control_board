import os
import re
from typing import Final
import openpyxl
import datetime
from login.login import _do_login
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.worksheet.hyperlink import Hyperlink
from selenium.webdriver.common.by import By
from time import sleep
from collections import defaultdict
from common_utils import normalize_string, set_cell_value, filter_unit_name_no_search_button, BOLD_FONT, CHECK_MARK, \
    LIGHT_BLUE_FILL, filter_unit_name_with_search_button, FamilyStatus, YELLOW_FILL
from flask import send_file, session
from io import BytesIO
import base64
import tempfile

# todo: move all constants to a separate constants file
URL_ACTIVE_TEAM_MEMBERS: Final[str] = 'https://app.paamonim.org.il/contacts/paam_index'
ULR_VACATION_TEAM_MEMBERS: Final[str] = 'https://app.paamonim.org.il/contacts/paam_index?in_vacation=1'
URL_FAMILIES_STATUS_PAGE: Final[str] = 'https://app.paamonim.org.il/budgets'
EXCEL_FILENAME: Final[str] = "cockpit.xlsx"
HEADER_NAME: Final[str] = "שם"
MAIN_SHEET_NAME: Final[str] = "ראשי"
FAMILIES_SHEET_NAME: Final[str] = "דוח משפחות"
MAIN_SHEET_FIRST_DATA_ROW_NUM: Final[int] = 6
FAMILIES_SHEET_FIRST_ROW_NUM: Final[int] = 3
HEADERS_ROW_NUM: Final[int] = 5

THIN_BORDER: Final[Border] = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                    bottom=Side(style='thin'))

thick_border_side = Side(style='thick')
no_border_side = Side(style=None)

right_top_border = Border(top=thick_border_side, left=no_border_side, right=thick_border_side, bottom=no_border_side)
left_top_border = Border(top=thick_border_side, left=thick_border_side, right=no_border_side, bottom=no_border_side)
right_bottom_border = Border(top=no_border_side, left=no_border_side, right=thick_border_side, bottom=thick_border_side)
left_bottom_border = Border(top=no_border_side, left=thick_border_side, right=no_border_side, bottom=thick_border_side)
top_border = Border(top=thick_border_side, left=no_border_side, right=no_border_side, bottom=no_border_side)
bottom_border = Border(top=no_border_side, left=no_border_side, right=no_border_side, bottom=thick_border_side)
left_border = Border(top=no_border_side, left=thick_border_side, right=no_border_side, bottom=no_border_side)
right_border = Border(top=no_border_side, left=no_border_side, right=thick_border_side, bottom=no_border_side)


# Get all header names
# header_names = [sheet.cell(row=HEADERS_ROW_NUM, column=col).value for col in range(1, sheet.max_column + 1) if sheet.cell(row=HEADERS_ROW_NUM, column=col).value is not None]
# print(f'### header_names: {header_names}')


def __find_header_index(sheet, header_name):
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=HEADERS_ROW_NUM, column=col).value
        if cell_value and normalize_string(cell_value) == normalize_string(header_name):
            return col

    print(f"Header '{header_name}' not found.")
    return None


def __apply_border_to_team_table(sheet, start_row, end_row, first_column_index, last_column_shift):
    sheet.cell(row=start_row, column=first_column_index).border = left_top_border
    sheet.cell(row=start_row, column=first_column_index + last_column_shift).border = right_top_border
    for column in range(first_column_index + 1, first_column_index + last_column_shift):
        sheet.cell(row=start_row, column=column).border = top_border
        sheet.cell(row=end_row, column=column).border = bottom_border

    sheet.cell(row=end_row, column=first_column_index).border = left_bottom_border
    sheet.cell(row=end_row, column=first_column_index + last_column_shift).border = right_bottom_border
    for row in range(start_row + 1, end_row):
        sheet.cell(row=row, column=first_column_index).border = left_border
        sheet.cell(row=row, column=first_column_index + last_column_shift).border = right_border


def update_wb_active_team_members(wb, sheet_name, start_row, header_name, team_list):
    # Select the active sheet
    sheet = wb[sheet_name]

    # Find the column index of the header
    column_index = __find_header_index(sheet, header_name)
    if column_index is None:
        return

    # Add new cells to the column
    i = start_row
    for key, values in team_list.items():
        team_lead_cell = sheet.cell(row=i, column=column_index)
        set_cell_value(team_lead_cell, key, LIGHT_BLUE_FILL)
        set_cell_value(sheet.cell(row=i, column=column_index + 1), CHECK_MARK)

        i += 1
        for value in values:
            set_cell_value(sheet.cell(row=i, column=column_index), value)
            set_cell_value(sheet.cell(row=i, column=column_index + 1), CHECK_MARK)
            i += 1

        i += 2  # Leave two empty rows to separate between team tables in the excel file


def retrieve_team_list(browser, unit_name, url_page, with_search_button=False, families_status=FamilyStatus.ACTIVE):
    # navigate to the urlpage
    browser.get(url_page)

    if with_search_button:
        filter_unit_name_with_search_button(browser, unit_name, families_status)
    else:
        filter_unit_name_no_search_button(browser, "user_", unit_name)

    team_list = defaultdict(set)
    rows = browser.find_elements(By.XPATH, './/tr[starts-with(@id, "user_")]')

    current_user = ""
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        current_user = cells[1].text if cells[1].text else current_user
        split_text = re.split('מרכז שרון - |מרכז שרון – ', cells[2].text)
        if len(split_text) > 1:
            team_list[split_text[1]].add(cells[1].text if cells[1].text else current_user)

    # filter out all branch unit entries and family pool entries 
    return {key: value for key, value in team_list.items() if "סניף" not in key and key != "מאגר משפחות לליווי"}


# todo: add column index instead of hard coded '3'
def collect_tutor_families(browser, unit_name, url_page, family_status):
    browser.get(url_page)

    filter_unit_name_with_search_button(browser, unit_name, family_status)

    active_families_list = defaultdict(lambda: [])
    rows = browser.find_elements(By.XPATH, './/tr[starts-with(@id, "family_")]')

    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        assigned_to = cells[3].text
        family_name = cells[0].text
        family_link = cells[0].find_element(By.TAG_NAME, "a").get_attribute("href")
        families = active_families_list[assigned_to]
        active_families_list[assigned_to] = families + [(family_name, family_link)]

    return active_families_list


def init_workbook(excel_filename):
    # copy the template file to the new excel file
    os.system(f'cp -f cockpit_template.xlsx {excel_filename}')

    # check if prev command ended successfully
    if not os.path.exists(excel_filename):
        print("Error copying the template file")
        exit(1)

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_filename)
    return wb


def save_workbook(wb):
    #wb.save(EXCEL_FILENAME)

    # excel_file = BytesIO()
    # wb.save(excel_file)
    # excel_file.seek(0)
    # # Encode the BytesIO object to a base64 string before storing it in the session
    # session['excel_file'] = base64.b64encode(excel_file.read()).decode('utf-8')


    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False)
    # Save the workbook to the temporary file
    wb.save(temp_file.name)
    # Store the temporary file's name in the session
    session['temp_file'] = temp_file.name


def __find_first_and_last_team_member_rows(sheet, start_row, team_leader, column_index):
    team_leader_row = None
    for row in range(start_row, sheet.max_row + 1):
        if sheet.cell(row=row, column=column_index).value == team_leader:
            team_leader_row = row
            break

    if team_leader_row is None:
        print(f"Error team leader '{team_leader}' not found.")
        return None, None

    # Find the last row of the team members under the team leader
    last_team_member_row = team_leader_row
    while sheet.cell(row=last_team_member_row + 1, column=column_index).value is not None:
        last_team_member_row += 1

    return team_leader_row, last_team_member_row


def update_wb_vacation_team_members(wb, sheet_name, header_name, start_row, team_list):
    # Select the active sheet
    sheet = wb[sheet_name]

    # Find the column index of the header
    column_index = __find_header_index(sheet, header_name)

    # Iterate over each team leader in the team_list
    for team_leader, team_members in team_list.items():
        # Find the first and last row of the team members under the team leader
        _, last_team_member_row = __find_first_and_last_team_member_rows(sheet, start_row, team_leader, column_index)

        # Insert the team members who are on vacation after the last row of the team members
        for i, team_member in enumerate(team_members, start=1):
            sheet.insert_rows(last_team_member_row + i)
            set_cell_value(sheet.cell(row=last_team_member_row + i, column=column_index), team_member)
            set_cell_value(sheet.cell(row=last_team_member_row + i, column=column_index + 2), CHECK_MARK)


def update_wb_families_status(wb, sheet_name, header_name, start_row, family_count_column_shift, family_list_column_shift, tutor_to_families):
    sheet = wb[sheet_name]

    # Find the column index of the header
    column_index = __find_header_index(sheet, header_name)

    # iterate over the rows in the header name index starting from row start_row
    for row in range(start_row, sheet.max_row + 1):
        # Check if the cell in the escort person column is valid for parsing
        if (sheet.cell(row=row, column=column_index).value and
                sheet.cell(row, column_index).fill.start_color.rgb == "00000000" and sheet.cell(row,
                                                                                                column_index).value != '-'):
            # Get the value of the escort person
            escort_person = sheet.cell(row=row, column=column_index).value
            # the escort families is a tuple of the number of families and a list of the families
            escort_families = tutor_to_families.get(escort_person, [])
            set_cell_value(sheet.cell(row=row, column=column_index + family_count_column_shift), len(escort_families))

            # Insert new rows for each additional family and create a hyperlink in each cell
            for i, (name, link) in enumerate(escort_families, start=1):
                if i > 1:
                    sheet.insert_rows(row + i - 1)
                    set_cell_value(sheet.cell(row + i - 1, column_index), '-')
                    set_cell_value(sheet.cell(row + i - 1, column_index + family_count_column_shift), '-')

                families_cell = sheet.cell(row=row + i - 1, column=column_index + family_list_column_shift)

                cell_value = f'=HYPERLINK("{link}", "{name}")'
                # cell_ref = f"{openpyxl.utils.get_column_letter(column_index+family_list_column_shift)}{row+i-1}"
                # hyperlink = Hyperlink(ref=cell_ref, display=name, location=link)
                print(f'### name: {name}  link:{link}')
                set_cell_value(families_cell, cell_value, font=Font(color=Color(rgb="0000FF"), underline='single'))
                # families_cell.hyperlink = hyperlink
                # families_cell.value = hyperlink.display

                families_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                # Adjust the height of the row to fit the content of the cell
                sheet.row_dimensions[
                    row + i - 1].height = 30  # Estimate the height based on the number of lines in the cell

    # Adjust the width of the column to be twice as wide as its current width
    column_letter = openpyxl.utils.get_column_letter(column_index + family_list_column_shift)
    sheet.column_dimensions[column_letter].width = sheet.column_dimensions[column_letter].width * 2


def create_families_sheet(wb, sheet_name, browser, start_row, tutor_to_families):
    sheet = wb[sheet_name]

    rows = browser.find_elements(By.XPATH, './/tr[starts-with(@id, "family_")]')

    i = start_row
    for (tutor, families) in tutor_to_families.items():
        # create a header line for this tutor
        sheet.merge_cells(start_row=i, start_column=6, end_row=i, end_column=19)
        merged_cell = sheet.cell(row=i, column=6)
        set_cell_value(merged_cell, tutor, fill=LIGHT_BLUE_FILL)
        i += 1

        # for each family of this tutor search for the family name in the html and copy relevant fields to excel
        for family in families:
            # find the row in the html that contains the family name
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if family[0] in cells[0].text:
                    # copy the row to the excel sheet
                    row_data = [cells[0].text, cells[1].text, cells[2].text, cells[12].text,
                                cells[13].text, cells[14].text, cells[7].text, cells[15].text, cells[9].text,
                                cells[11].text, cells[10].text, cells[8].text, cells[6].text]
                    # write row_data to the excel at row
                    for col, cell in enumerate(row_data, start=6):
                        set_cell_value(sheet.cell(row=i, column=col), cell)
                        # Adjust the width of the column to text length #todo: put this adjustment into a function
                        column_letter = openpyxl.utils.get_column_letter(col)
                        if len(sheet.cell(i, col).value) > sheet.column_dimensions[column_letter].width:
                            sheet.column_dimensions[column_letter].width = len(sheet.cell(i, col).value)
                    num_skip_lines = write_family_alerts(cells, sheet, i)
                    i += num_skip_lines if num_skip_lines > 0 else 1
                    break
    return i


def write_family_alerts(cells, sheet, row):
    alerts = []
    if not cells[8].text and int(cells[6].text.split()[0]) > 45:
        alerts.append("ליווי בן יותר מ-45 יום ועדיין ללא תקציב ")
    if not cells[12].text.strip():
        alerts.append("אין פגישה אחרונה בתיק")
    else:
        # parse the date string
        last_meeting_date = datetime.datetime.strptime(cells[12].text.strip(), "%d-%m-%y")
        # get the current date
        current_date = datetime.datetime.now()
        # get the current month and the previous month
        current_month = current_date.month
        previous_month = current_month - 1 if current_month != 1 else 12
        # if the last meeting date's month is not the same as the current month or the previous month, print the alert
        if last_meeting_date.month != current_month and last_meeting_date.month != previous_month:
            alerts.append(
                f'לא התקיימה פגישה בחודש הנוכחי או הקודם')
    if not cells[13].text.strip():
        alerts.append("לא נקבעה הפגישה הבאה")

    for i, alert in enumerate(alerts, start=1):
        if i > 1:
            sheet.insert_rows(row + i - 1)
        set_cell_value(sheet.cell(row=row + i - 1, column=19), alert, fill=YELLOW_FILL)
        # Adjust the width of the column to text length #todo: put this adjustment into a function
        column_letter = openpyxl.utils.get_column_letter(19)
        if len(sheet.cell(row + i - 1, 19).value) > sheet.column_dimensions[column_letter].width:
            sheet.column_dimensions[column_letter].width = len(sheet.cell(row+i-1, 19).value)

    print(f'### alerts for family {cells[0].text}: {alerts}')
    return len(alerts)


def apply_borders_to_all_teams(wb, sheet_name, header_name, start_row, team_list):
    sheet = wb[sheet_name]

    # Find the column index of the header
    column_index = __find_header_index(sheet, header_name)

    # Iterate over each team leader in the all_teams
    for team_leader in team_list.keys():
        # Find the first and last row of the team members under the team leader
        team_leader_row, last_team_member_row = __find_first_and_last_team_member_rows(sheet, start_row, team_leader,
                                                                                       column_index)

        # Apply border to the team table
        __apply_border_to_team_table(sheet, team_leader_row, last_team_member_row, column_index, 7)


def insert_totals(wb, sheet_name, start_row, header_name):
    # Select the active sheet
    sheet = wb[sheet_name]

    # Find the column index of the header
    column_index = __find_header_index(sheet, header_name)

    # Initialize counters
    total_counter = 0
    active_counter = 0
    vacation_counter = 0
    blank_rows_count = 0

    # Iterate over each row starting from the start row
    for row in range(start_row, sheet.max_row + 1):
        # Check if the cell in the active column and the vacation column has a check mark
        if sheet.cell(row=row, column=column_index).value is not None:
            total_counter += 1
            if sheet.cell(row=row, column=column_index + 1).value == u'\u2714':
                active_counter += 1
            if sheet.cell(row=row, column=column_index + 2).value == u'\u2714':
                vacation_counter += 1
            blank_rows_count = 0
        else:
            if blank_rows_count == 0:
                set_cell_value(sheet.cell(row=row, column=column_index), total_counter, LIGHT_BLUE_FILL)
                set_cell_value(sheet.cell(row=row, column=column_index + 1), active_counter, LIGHT_BLUE_FILL)
                set_cell_value(sheet.cell(row=row, column=column_index + 2), vacation_counter, LIGHT_BLUE_FILL)
                total_counter = 0
                active_counter = 0
                vacation_counter = 0
            blank_rows_count += 1
            if blank_rows_count == 2:
                row += 2
            elif blank_rows_count > 2:
                break

    # Insert a new row at the end
    sheet.append([])


def main(browser, unit_name):
    # app = QApplication([])
    # browser, unit_name = _do_login()
    if not browser:
        print("error occurred. exiting gracefully")
        exit(0)

    wb = init_workbook(EXCEL_FILENAME)

    active_team_list = retrieve_team_list(browser, unit_name, URL_ACTIVE_TEAM_MEMBERS)
    print(f'active team list: {active_team_list}')

    # add active team members to the excel file
    update_wb_active_team_members(wb, MAIN_SHEET_NAME, MAIN_SHEET_FIRST_DATA_ROW_NUM, HEADER_NAME, active_team_list)

    vacation_team_list = retrieve_team_list(browser, unit_name, ULR_VACATION_TEAM_MEMBERS)
    print(f'vacation team list: {vacation_team_list}')

    # add vacation team members to the excel file
    update_wb_vacation_team_members(wb, MAIN_SHEET_NAME, HEADER_NAME, MAIN_SHEET_FIRST_DATA_ROW_NUM, vacation_team_list)

    tutor_to_families = collect_tutor_families(browser, unit_name, URL_FAMILIES_STATUS_PAGE, FamilyStatus.ACTIVE)
    print(f'active families list: {tutor_to_families}')

    # add the amount of active families + links (per tutor) to the excel file
    update_wb_families_status(wb, MAIN_SHEET_NAME, HEADER_NAME, MAIN_SHEET_FIRST_DATA_ROW_NUM,
                              5, 6, tutor_to_families)

    tutor_to_ready_families = collect_tutor_families(browser, unit_name, URL_FAMILIES_STATUS_PAGE, FamilyStatus.READY_TO_START)
    print(f'ready to start families list: {tutor_to_ready_families}')

    update_wb_families_status(wb, MAIN_SHEET_NAME, HEADER_NAME, MAIN_SHEET_FIRST_DATA_ROW_NUM, 3, 4, tutor_to_ready_families)

    insert_totals(wb, MAIN_SHEET_NAME, MAIN_SHEET_FIRST_DATA_ROW_NUM, HEADER_NAME)

    # apply borders to all team tables
    apply_borders_to_all_teams(wb, MAIN_SHEET_NAME, HEADER_NAME, MAIN_SHEET_FIRST_DATA_ROW_NUM,
                               {**active_team_list, **vacation_team_list})

    # to reset the checkboxes checked by previous steps
    browser.get(URL_FAMILIES_STATUS_PAGE)
    filter_unit_name_with_search_button(browser, unit_name)

    num_of_table_rows = create_families_sheet(wb, FAMILIES_SHEET_NAME, browser, FAMILIES_SHEET_FIRST_ROW_NUM, tutor_to_families)
    __apply_border_to_team_table(wb[FAMILIES_SHEET_NAME], 1, num_of_table_rows-1, 6, 13)

    save_workbook(wb)
    print(f'### DONE')


# if __name__ == "__main__":
#     main()
