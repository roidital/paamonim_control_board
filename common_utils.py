from typing import Final

from PyQt5.QtWidgets import QApplication, QMessageBox
from openpyxl.styles import Alignment
import unicodedata
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl.styles import Font, PatternFill
from enum import Enum

CHECK_MARK = u'\u2714'
BOLD_FONT: Final[Font] = Font(bold=True, size=14)
LIGHT_BLUE_FILL: Final[PatternFill] = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
app = QApplication([])  # QApplication instance is required for QMessageBox


# enum for family status
class FamilyStatus(Enum):
    ACTIVE = 1
    READY_TO_START = 2
    ENDED = 3


def normalize_string(s):
    """
    Normalize a string by removing diacritics and converting to lowercase.
    """
    return unicodedata.normalize("NFD", s).casefold()


def set_cell_value(cell, value, fill=None, font=BOLD_FONT):
    cell.value = value
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')


def filter_unit_name_no_search_button(browser, filter_by, unit_name):
    # wait for the dropdown to be clickable
    dropdown = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.CLASS_NAME, 'betterselecter-sel')))
    dropdown.click()

    # wait for the options to be visible
    WebDriverWait(browser, 5).until(EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="betterselecter-op"]')))

    # count the rows before filtering the table
    current_row_count = len(browser.find_elements(By.XPATH, f'.//tr[starts-with(@id, {filter_by})]'))

    try:
        option = WebDriverWait(browser, 5).until(EC.visibility_of_element_located((By.XPATH, f'//div[@class="betterselecter-op" and contains(text(), "{unit_name}")]')))
        print("### Found the option")
        option.click()
    except TimeoutException:
        QMessageBox.information(None, "שגיאה", "לא נמצאה יחידה עם שם זה")
        exit(0)


    def __rows_have_updated(browser):
        new_row_count = len(browser.find_elements(By.XPATH, f'.//tr[starts-with(@id, {filter_by})]'))
        return new_row_count != current_row_count

    # wait for the table to be updated after unit filter
    WebDriverWait(browser, 10).until(__rows_have_updated)


def filter_unit_name_with_search_button(browser, unit_name, families_status = FamilyStatus.ACTIVE):
    # wait for the dropdown to be clickable
    dropdown = WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.CLASS_NAME, 'betterselecter-sel')))
    dropdown.click()

    # wait for the options to be visible
    options = WebDriverWait(browser, 5).until(
        EC.visibility_of_all_elements_located((By.XPATH, '//div[@class="betterselecter-op"]')))

    # find the relevant unit we wish to analyze
    found_option = False
    for option in options:
        if unit_name in option.text:
            print(f"### Found the option {unit_name}")
            option.click()
            found_option = True
            break

    if not found_option:
        QMessageBox.information(None, "שגיאה", "לא נמצאה יחידה עם שם זה")
        exit(0)

    if families_status == FamilyStatus.READY_TO_START:
        browser.find_element(By.ID, 'started').click()
        browser.find_element(By.ID, 'ordered').click()

    browser.find_element(By.ID, 'searchButton').click()

    # find the table
    tables = WebDriverWait(browser, 5).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'tbl_chart')))

    # select the second table
    table = tables[1]
    return table
