from openpyxl.styles import Alignment
import unicodedata
import openpyxl
from src.common.constants import LEFT_TOP_BORDER, RIGHT_TOP_BORDER, TOP_BORDER, BOTTOM_BORDER, LEFT_BOTTOM_BORDER, \
    RIGHT_BOTTOM_BORDER, LEFT_BORDER, RIGHT_BORDER, BOLD_FONT, FamilyStatus, LIGHT_BLUE_FILL
import asyncio


# app = QApplication([])  # QApplication instance is required for QMessageBox


def normalize_string(s):
    """
    Normalize a string by removing diacritics and converting to lowercase.
    """
    return unicodedata.normalize("NFD", s).casefold()


def set_cell_value(cell, value, fill=None, font=BOLD_FONT, adjust_width=False, wrap_text=False):
    cell.value = value
    cell.font = font
    if fill:
        cell.fill = fill
    if wrap_text:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if adjust_width:
        __adjust_column_width_to_text(cell)


async def filter_unit_name_no_search_button(page, unit_name):
    # wait for the dropdown to be clickable
    dropdown = await page.waitForSelector('.betterselecter-sel', {'visible': True})
    await dropdown.click()

    # wait for the options to be visible
    await page.waitForSelector('.betterselecter-op', {'visible': True})

    # count the rows before filtering the table
    #current_row_count = len(await page.querySelectorAll(f'tr[id^={filter_by}]'))
    try:
        option = await page.waitForXPath(f'//div[@class="betterselecter-op" and contains(text(), "{unit_name}")]',
                                         {'visible': True})
        print("### Found the option")
        await option.click()
    except:
        print(f'### ERROR: unit name: {unit_name} not found')
        exit(0)

    # while True:
    #     new_row_count = len(await page.querySelectorAll(f'tr[id^={filter_by}]'))
    #     if new_row_count != current_row_count:
    #         break
    #     await asyncio.sleep(0.1)  # wait a bit before checking again
    await asyncio.sleep(2)


async def filter_unit_name_with_search_button(page, unit_name, families_status = FamilyStatus.ACTIVE):
    # wait for the dropdown to be clickable
    dropdown = await page.waitForSelector('.betterselecter-sel', {'visible': True})
    await dropdown.click()

    # wait for the options to be visible
    await page.waitForSelector('.betterselecter-op', {'visible': True})

    # find the relevant unit we wish to analyze
    try:
        option = await page.waitForXPath(f'//div[@class="betterselecter-op" and contains(text(), "{unit_name}")]',
                                         {'visible': True})
        print("### Found the option")
        await option.click()
    except:
        print(f'### ERROR: unit name: {unit_name} not found')
        exit(0)

    if families_status == FamilyStatus.READY_TO_START:
        started_filter = await page.querySelector('#started')
        await started_filter.click()
        ordered_filter = await page.querySelector('#ordered')
        await ordered_filter.click()

    search_button = await page.querySelector('#searchButton')
    await search_button.click()

    # find the table
    table = await page.waitForSelector('.tbl_chart:nth-of-type(2)')

    # return the second table
    return table


def set_sum_formula_to_cell(sheet, start_row, end_row, column_index, divide_by_2=False):
    column_letter = openpyxl.utils.get_column_letter(column_index)
    formula = f'=SUM({column_letter}{start_row}:{column_letter}{end_row - 1})'
    set_cell_value(sheet.cell(row=end_row, column=column_index), (formula + '/2' if divide_by_2 else formula), LIGHT_BLUE_FILL)


def __apply_border_to_team_table(sheet, start_row, end_row, first_column_index, last_column_shift):
    sheet.cell(row=start_row, column=first_column_index).border = LEFT_TOP_BORDER
    sheet.cell(row=start_row, column=first_column_index + last_column_shift).border = RIGHT_TOP_BORDER
    for column in range(first_column_index + 1, first_column_index + last_column_shift):
        sheet.cell(row=start_row, column=column).border = TOP_BORDER
        sheet.cell(row=end_row, column=column).border = BOTTOM_BORDER

    sheet.cell(row=end_row, column=first_column_index).border = LEFT_BOTTOM_BORDER
    sheet.cell(row=end_row, column=first_column_index + last_column_shift).border = RIGHT_BOTTOM_BORDER
    for row in range(start_row + 1, end_row):
        sheet.cell(row=row, column=first_column_index).border = LEFT_BORDER
        sheet.cell(row=row, column=first_column_index + last_column_shift).border = RIGHT_BORDER


# def __find_header_index(sheet, header_name):
#     for col in range(1, sheet.max_column + 1):
#         cell_value = sheet.cell(row=HEADERS_ROW_NUM, column=col).value
#         if cell_value and normalize_string(cell_value) == normalize_string(header_name):
#             return col
#
#     print(f"Header '{header_name}' not found.")
#     return None


def __adjust_column_width_to_text(cell):
    column_letter = openpyxl.utils.get_column_letter(cell.column)
    cell_value_str = str(cell.value)
    if len(cell_value_str) > cell.parent.column_dimensions[column_letter].width:
        cell.parent.column_dimensions[column_letter].width = len(cell_value_str) * 1.1


def __adjust_row_height_to_text(cell):
    cell_value_str = str(cell.value)
    if len(cell_value_str) > cell.parent.row_dimensions[cell.row].height:
        cell.parent.row_dimensions[cell.row].height = len(cell_value_str) * 1.1
