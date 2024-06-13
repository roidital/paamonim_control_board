from typing import Final
from openpyxl.styles import Border, Side, Font, PatternFill
from enum import Enum


# URLs
MAIN_LOGIN_URL: Final[str] = 'https://app.paamonim.org.il'
BUDGET_AND_BALANCES_PAGE: Final[str] = 'https://app.paamonim.org.il/budgets/budget_and_balances/'
URL_ACTIVE_TEAM_MEMBERS: Final[str] = 'https://app.paamonim.org.il/contacts/paam_index'
ULR_VACATION_TEAM_MEMBERS: Final[str] = 'https://app.paamonim.org.il/contacts/paam_index?in_vacation=1'
URL_FAMILIES_STATUS_PAGE: Final[str] = 'https://app.paamonim.org.il/budgets'

# Excel constants
EXCEL_FILENAME: Final[str] = "../cockpit.xlsx"
HEADER_NAME: Final[str] = "שם"
TEAMS_LIST_SHEET_NAME: Final[str] = "צוותים"
FAMILIES_SHEET_NAME: Final[str] = "דוח משפחות"

# column//row indices
TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM: Final[int] = 6
FAMILIES_SHEET_FIRST_ROW_NUM: Final[int] = 4
HEADERS_ROW_NUM: Final[int] = 5
TUTOR_COLUMN_IN_TEAMS_SHEET: Final[int] = 3
FAMILIES_SHEET_FIRST_COLUMN_INDEX: Final[int] = 2
FAMILIES_SHEET_LAST_COLUMN_INDEX: Final[int] = 22
DAYS_WITHOUT_BUDGET_LIMIT: Final[int] = 45
DAYS_WITHOUT_FIRST_MEETING_LIMIT: Final[int] = 30
READY_FAMILIES_SUM_COLUMN_DIFF: Final[int] = 3
ACTIVE_FAMILIES_SUM_COLUMN_DIFF: Final[int] = 5

# family data fields names
FAMILY_NAME: Final[str] = "family_name"
UNIT_NAME: Final[str] = "unit_name"
CITY: Final[str] = "city"
TUTOR: Final[str] = "tutor"
LAST_MEETING_DATE: Final[str] = "last_meeting_date"
NEXT_MEETING_DATE: Final[str] = "next_meeting_date"
LAST_SHIKUF_BITSUA: Final[str] = "last_shikuf_bitsua"
LAST_OSH_STATS: Final[str] = "last_osh_stats"
TOTAL_DEBTS: Final[str] = "total_debts"
MONTHLY_DEBTS_PAYMENT: Final[str] = "monthly_debts_payment"
UNSETTLED_DEBTS: Final[str] = "unsettled_debts"
BUDGET: Final[str] = "budget"
CASE_AGE: Final[str] = "case_age"
NUM_OF_MEETINGS: Final[str] = "num_of_meetings"
NUM_CANCELLED_MEETINGS: Final[str] = "num_cancelled_meetings"
BUDGET_INCOME: Final[str] = "budget_income"
BUDGET_EXPENSE: Final[str] = "budget_expense"
BUDGET_DIFF: Final[str] = "budget_diff"
MONTH_INCOME: Final[str] = "month_income"
MONTH_EXPENSE: Final[str] = "month_expense"
LAST_MONTH_DIFF: Final[str] = "last_month_diff"


# borders
THIN_BORDER: Final[Border] = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                    bottom=Side(style='thin'))
THICK_BORDER_SIDE = Side(style='thick')
NO_BORDER_SIDE = Side(style=None)
RIGHT_TOP_BORDER = Border(top=THICK_BORDER_SIDE, left=NO_BORDER_SIDE, right=THICK_BORDER_SIDE, bottom=NO_BORDER_SIDE)
LEFT_TOP_BORDER = Border(top=THICK_BORDER_SIDE, left=THICK_BORDER_SIDE, right=NO_BORDER_SIDE, bottom=NO_BORDER_SIDE)
RIGHT_BOTTOM_BORDER = Border(top=NO_BORDER_SIDE, left=NO_BORDER_SIDE, right=THICK_BORDER_SIDE, bottom=THICK_BORDER_SIDE)
LEFT_BOTTOM_BORDER = Border(top=NO_BORDER_SIDE, left=THICK_BORDER_SIDE, right=NO_BORDER_SIDE, bottom=THICK_BORDER_SIDE)
TOP_BORDER = Border(top=THICK_BORDER_SIDE, left=NO_BORDER_SIDE, right=NO_BORDER_SIDE, bottom=NO_BORDER_SIDE)
BOTTOM_BORDER = Border(top=NO_BORDER_SIDE, left=NO_BORDER_SIDE, right=NO_BORDER_SIDE, bottom=THICK_BORDER_SIDE)
LEFT_BORDER = Border(top=NO_BORDER_SIDE, left=THICK_BORDER_SIDE, right=NO_BORDER_SIDE, bottom=NO_BORDER_SIDE)
RIGHT_BORDER = Border(top=NO_BORDER_SIDE, left=NO_BORDER_SIDE, right=THICK_BORDER_SIDE, bottom=NO_BORDER_SIDE)

# colors and fonts
CHECK_MARK = u'\u2714'
BOLD_FONT: Final[Font] = Font(bold=True, size=14)
LIGHT_BLUE_FILL: Final[PatternFill] = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
YELLOW_FILL: Final[PatternFill] = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# enums
class FamilyStatus(Enum):
    ACTIVE = 1
    READY_TO_START = 2
    ENDED = 3
