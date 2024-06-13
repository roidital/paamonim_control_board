import re
import openpyxl
import datetime
from src.common.common_utils import filter_unit_name_with_search_button, set_cell_value, __apply_border_to_team_table, \
    __adjust_column_width_to_text
from src.common.constants import URL_FAMILIES_STATUS_PAGE, LIGHT_BLUE_FILL, FAMILIES_SHEET_NAME, YELLOW_FILL, \
    FAMILIES_SHEET_FIRST_COLUMN_INDEX, FAMILIES_SHEET_LAST_COLUMN_INDEX, DAYS_WITHOUT_BUDGET_LIMIT, MAIN_LOGIN_URL, \
    BUDGET_AND_BALANCES_PAGE, FAMILY_NAME, UNIT_NAME, CITY, LAST_MEETING_DATE, NEXT_MEETING_DATE, LAST_SHIKUF_BITSUA, \
    LAST_OSH_STATS, TOTAL_DEBTS, MONTHLY_DEBTS_PAYMENT, UNSETTLED_DEBTS, BUDGET, CASE_AGE, NUM_OF_MEETINGS, \
    NUM_CANCELLED_MEETINGS, BUDGET_INCOME, BUDGET_EXPENSE, BUDGET_DIFF, MONTH_INCOME, MONTH_EXPENSE, LAST_MONTH_DIFF, \
    TUTOR, DAYS_WITHOUT_FIRST_MEETING_LIMIT
from selenium.webdriver.common.by import By
from collections import defaultdict
import asyncio
from pyppeteer import launch
import nest_asyncio
nest_asyncio.apply()


def create_families_sheet(wb, sheet_name, browser, start_row, tutor_to_families, unit_name, username, password):
    # to reset the checkboxes checked by previous steps
    browser.get(URL_FAMILIES_STATUS_PAGE)
    filter_unit_name_with_search_button(browser, unit_name)

    sheet = wb[sheet_name]

    rows = browser.find_elements(By.XPATH, './/tr[starts-with(@id, "family_")]')

    i = start_row
    family_data_dict = defaultdict(lambda: [])
    for (tutor, families) in tutor_to_families.items():
        # for each family of this tutor search for the family name in the html and copy relevant fields to excel
        for family in families:
            # find the row in the html that contains the family name
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if family[0] in cells[0].text:
                    # get the the family's id number (from html)
                    family_id = row.get_attribute('id').split('_')[1]  # Get the id attribute
                    retrieve_data_from_common_families_table(row, family_id, family_data_dict)
                    family_data_dict[family_id]['line_num'] = i
                    set_values_from_common_families_table_to_excel(family_data_dict[family_id], sheet)

                    write_family_alerts(family_data_dict[family_id], sheet, i)
                    i += 1
                    break
    # retrieve the budget and balances data for each family (parallel execution)
    asyncio.run(browser_dispatcher(family_data_dict, username, password))
    # print(f'### AFTER family_data_dict: {family_data_dict}')

    for family_id in family_data_dict.keys():
        set_budget_and_balances_to_excel(family_data_dict[family_id], sheet)

    num_of_table_rows = i
    __apply_border_to_team_table(wb[FAMILIES_SHEET_NAME], 1, num_of_table_rows - 1,
                                 FAMILIES_SHEET_FIRST_COLUMN_INDEX,
                                 (FAMILIES_SHEET_LAST_COLUMN_INDEX-FAMILIES_SHEET_FIRST_COLUMN_INDEX))


# family_data_dict is an output parameter, a dictionary populated families data
def retrieve_data_from_common_families_table(row, family_id, family_data_dict):
    cells = row.find_elements(By.TAG_NAME, "td")
    if cells[0].text:
        # todo: consider moving the keys in the inner dict to constants in constants.py
        family_data_dict[family_id] = {FAMILY_NAME: cells[0].text,
                                       UNIT_NAME: cells[1].text,
                                       CITY: cells[2].text,
                                       TUTOR: cells[3].text,
                                       LAST_MEETING_DATE: cells[12].text,
                                       NEXT_MEETING_DATE: cells[13].text,
                                       LAST_SHIKUF_BITSUA: cells[7].text,
                                       LAST_OSH_STATS: cells[15].text,
                                       TOTAL_DEBTS: cells[9].text,
                                       MONTHLY_DEBTS_PAYMENT: cells[11].text,
                                       UNSETTLED_DEBTS: cells[10].text,
                                       BUDGET: cells[8].text,
                                       CASE_AGE: cells[6].text}
        # parse the number of meetings and the number of cancelled meetings from the format "num_meetings (num_cancelled_meetings)"
        match = re.match(r"(\d+) \((\d+)\)", cells[14].text)
        if match:
            family_data_dict[family_id][NUM_OF_MEETINGS] = match.group(1)
            family_data_dict[family_id][NUM_CANCELLED_MEETINGS] = match.group(2)
        else:
            family_data_dict[family_id][NUM_OF_MEETINGS] = cells[14].text


def set_values_from_common_families_table_to_excel(family_data, sheet):
    row = family_data['line_num']
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX),
                   family_data[UNIT_NAME], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 1),
                   family_data[TUTOR], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 2),
                   family_data[FAMILY_NAME], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 3),
                   family_data[CITY], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 4),
                   family_data[CASE_AGE], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 5),
                   family_data[LAST_MEETING_DATE], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 6),
                   family_data[NEXT_MEETING_DATE], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 7),
                   family_data[NUM_OF_MEETINGS], adjust_width=True)
    if NUM_CANCELLED_MEETINGS in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 8),
                       family_data[NUM_CANCELLED_MEETINGS], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 18),
                   family_data[LAST_OSH_STATS], adjust_width=True)


def set_budget_and_balances_to_excel(family_data, sheet):
    row = family_data['line_num']
    if BUDGET_INCOME in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 9),
                       family_data[BUDGET_INCOME], adjust_width=True)
    if BUDGET_EXPENSE in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 10),
                       family_data[BUDGET_EXPENSE], adjust_width=True)
    if BUDGET_DIFF in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 11),
                       family_data[BUDGET_DIFF], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 12),
                   family_data[TOTAL_DEBTS], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 13),
                   family_data[MONTHLY_DEBTS_PAYMENT], adjust_width=True)
    set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 14),
                   family_data[UNSETTLED_DEBTS], adjust_width=True)
    if MONTH_INCOME in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 15),
                       family_data[MONTH_INCOME], adjust_width=True)
    if MONTH_EXPENSE in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 16),
                       family_data[MONTH_EXPENSE], adjust_width=True)
    if LAST_MONTH_DIFF in family_data:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_FIRST_COLUMN_INDEX + 17),
                       family_data[LAST_MONTH_DIFF], adjust_width=True)


async def auto_login(username, password):
    options = {
        'ignoreHTTPSErrors': True,
        'args': ['--no-sandbox'],
        # since we are using Flask to launch this webapp - this flow is not running in main thread, in python signals
        # can be set only in main thread, so we need to disable the signals handling
        'handleSIGINT': False,
        'handleSIGTERM': False,
        'handleSIGHUP': False
    }
    browser = await launch(options=options)
    page = await browser.newPage()

    # navigate to the login page
    await page.goto(MAIN_LOGIN_URL)

    # Perform login, since these are the same username and password used for the initial login, we are not supposed
    # to reach here unless the login was successful, however it's better to check for a failed login and handle it (although not expected)
    # todo: add a check for a failed login
    await page.type('input[name=login]', username)
    await page.type('input[name=password]', password)
    await page.click('#loginBtn')

    # Wait for navigation to complete
    await page.waitForNavigation()

    return browser


async def browser_dispatcher(family_data_dict, username, password):
    # Perform login
    browser = await auto_login(username, password)

    # in case the family doesn't have a shikuf/bitsua - it means the BUDGET_AND_BALANCES_PAGE page doesn't exist
    # which will follow a timeout exception in the fetch_family_data function
    tasks = [fetch_family_data(browser, family_id, family_data_dict) for family_id in family_data_dict.keys()
             if family_data_dict[family_id]['last_shikuf_bitsua'] != '']
    pages_content = await asyncio.gather(*tasks)

    for page_content in pages_content:
        print(f'### page_content: {page_content}')

    await browser.close()


async def fetch_family_data(browser, family_id, family_data_dict):
    page = await browser.newPage()
    await page.goto(BUDGET_AND_BALANCES_PAGE + family_id)

    try:
        await page.waitForSelector('#expenseTable')
    except:
        print(f'### ERROR - family {family_id} got timed out while waiting for #expenseTable')
        await page.close()
        return 'done with family_id: ' + family_id

    budget_income = await page.evaluate('''() => {
            const td = document.querySelector('#sumTable tr.incomeTitle td.sumLine1Col1');
            return td ? td.innerHTML : null;
        }''')
    # print(f'### income: {budget_income}')
    family_data_dict[family_id][BUDGET_INCOME] = budget_income
    budget_expense = await page.evaluate('''() => {
                const td = document.querySelector('#sumTable tr.expenseTitle td.sumLine2Col1');
                return td ? td.innerHTML : null;
            }''')
    # print(f'### expense: {budget_expense}')
    family_data_dict[family_id][BUDGET_EXPENSE] = budget_expense
    if budget_income and budget_expense:
        # print(f'### budget diff: {int(budget_income) - int(budget_expense)}')
        family_data_dict[family_id][BUDGET_DIFF] = int(budget_income) - int(budget_expense)
    month_income = await page.evaluate('''() => {
                    const td = document.querySelector('#sumTable tr.incomeTitle td.sumLine1Col3');
                    return td ? td.innerHTML : null;
                }''')
    # print(f'### last month income: {month_income}')
    family_data_dict[family_id][MONTH_INCOME] = month_income
    month_expense = await page.evaluate('''() => {
                        const td = document.querySelector('#sumTable tr.expenseTitle td.sumLine2Col3');
                        return td ? td.innerHTML : null;
                    }''')
    # print(f'### last month expense: {month_expense}')
    family_data_dict[family_id][MONTH_EXPENSE] = month_expense
    if month_income and month_expense:
        # print(f'### last month diff: {int(month_income) - int(month_expense)}')
        family_data_dict[family_id][LAST_MONTH_DIFF] = int(month_income) - int(month_expense)

    await page.close()
    return 'done with family_id: ' + family_id


def write_family_alerts(family_data, sheet, row):
    print(f'### alerts. cells: {family_data}')
    alerts = []
    if not family_data[BUDGET] and int(family_data[CASE_AGE].split()[0]) > DAYS_WITHOUT_BUDGET_LIMIT:
        alerts.append("ליווי בן יותר מ-45 יום ועדיין ללא תקציב ")
    if not family_data[LAST_MEETING_DATE].strip():
        if int(family_data[CASE_AGE].split()[0]) > DAYS_WITHOUT_FIRST_MEETING_LIMIT:
            alerts.append("אין פגישה אחרונה בתיק שמתנהל מעל 30 יום")
    else:
        # parse the date string
        last_meeting_date = datetime.datetime.strptime(family_data[LAST_MEETING_DATE].strip(), "%d-%m-%y")
        current_date = datetime.datetime.now()
        current_month = current_date.month
        previous_month = current_month - 1 if current_month != 1 else 12
        # if the last meeting date's month is not the same as the current month or the previous month, add an alert
        if last_meeting_date.month != current_month and last_meeting_date.month != previous_month:
            alerts.append(
                f'לא התקיימה פגישה בחודש הנוכחי או הקודם')
    if not family_data[NEXT_MEETING_DATE].strip():
        alerts.append("לא נקבעה הפגישה הבאה")
    if MONTH_INCOME in family_data and BUDGET_INCOME in family_data:
        if family_data[MONTH_INCOME] < float(family_data[BUDGET_INCOME]*0.75):
            alerts.append("הכנסה חודשית נמוכה מ-75% מהתקציב החודשי")
    if MONTH_EXPENSE in family_data and BUDGET_EXPENSE in family_data:
        if family_data[MONTH_EXPENSE] > float(family_data[BUDGET_EXPENSE]*1.3):
            alerts.append("הוצאה חודשית גבוהה ביותר מ-30% מהתקציב החודשי")

    # concat all the alerts into one string with a new line separator
    alerts = '\n'.join(alerts)
    if alerts:
        set_cell_value(sheet.cell(row=row, column=FAMILIES_SHEET_LAST_COLUMN_INDEX), alerts, fill=YELLOW_FILL, adjust_width=True, wrap_text=True)

    print(f'### alerts for family {family_data[FAMILY_NAME]}: {alerts}')
