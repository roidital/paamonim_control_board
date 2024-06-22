import re
import openpyxl
from src.common.common_utils import set_cell_value, filter_unit_name_with_search_button, \
    filter_unit_name_no_search_button, __apply_border_to_team_table, set_sum_formula_to_cell
from src.common.constants import URL_ACTIVE_TEAM_MEMBERS, TEAMS_LIST_SHEET_NAME, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM, \
    ULR_VACATION_TEAM_MEMBERS, URL_FAMILIES_STATUS_PAGE, CHECK_MARK, FamilyStatus, LIGHT_BLUE_FILL, \
    TUTOR_COLUMN_IN_TEAMS_SHEET, READY_FAMILIES_SUM_COLUMN_DIFF, ACTIVE_FAMILIES_SUM_COLUMN_DIFF, \
    TEAMS_SHEET_NAME_HEADER_COLUMN_INDEX, ACTIVE_FAMILY_COUNT_COLUMN_SHIFT, ACTIVE_FAMILY_LIST_COLUMN_SHIFT, \
    READY_FAMILY_COUNT_COLUMN_SHIFT, READY_FAMILY_LIST_COLUMN_SHIFT
from collections import defaultdict
from openpyxl.styles import Font, Alignment, Color


async def create_teams_list_sheet(browser, unit_name, wb):
    active_team_list = await retrieve_team_list(browser, unit_name, URL_ACTIVE_TEAM_MEMBERS)
    print(f'active team list: {active_team_list}')

    sheet = wb[TEAMS_LIST_SHEET_NAME]
    # add active team members to the excel file
    update_wb_active_team_members(sheet, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM, active_team_list)

    vacation_team_list = await retrieve_team_list(browser, unit_name, ULR_VACATION_TEAM_MEMBERS)
    print(f'vacation team list: {vacation_team_list}')

    # add vacation team members to the excel file
    update_wb_vacation_team_members(sheet, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM, vacation_team_list)

    tutor_to_families, team_leader_to_families = await collect_tutor_families(browser, unit_name, URL_FAMILIES_STATUS_PAGE,
                                                                        FamilyStatus.ACTIVE)
    # print(f'team_leader_to_families: {team_leader_to_families}')

    # add the amount of active families + links (per tutor) to the excel file
    update_wb_families_status(sheet, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM, ACTIVE_FAMILY_COUNT_COLUMN_SHIFT,
                              ACTIVE_FAMILY_LIST_COLUMN_SHIFT, tutor_to_families)

    tutor_to_ready_families, _ = await collect_tutor_families(browser, unit_name, URL_FAMILIES_STATUS_PAGE,
                                                        FamilyStatus.READY_TO_START)
    print(f'ready to start families list: {tutor_to_ready_families}')

    update_wb_families_status(sheet, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM, READY_FAMILY_COUNT_COLUMN_SHIFT,
                              READY_FAMILY_LIST_COLUMN_SHIFT, tutor_to_ready_families)

    insert_totals(sheet, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM)

    # apply borders to all team tables
    apply_borders_to_all_teams(sheet, TEAM_LISTS_SHEET_FIRST_DATA_ROW_NUM,{**active_team_list, **vacation_team_list})
    # return tutor_to_families since it's required in create_families_sheet() called from main
    return team_leader_to_families


async def retrieve_team_list(browser, unit_name, url_page, with_search_button=False, families_status=FamilyStatus.ACTIVE):
    page = await browser.newPage()
    await page.goto(url_page)

    # roi: I'm here - use page instead of browser to find elements
    if with_search_button:
        await filter_unit_name_with_search_button(browser, unit_name, families_status)
    else:
        await filter_unit_name_no_search_button(page, unit_name)
        print('### filter_unit_name_no_search_button DONE')

    team_list = defaultdict(set)
    rows = await page.querySelectorAll('tr[id^="user_"]')

    current_user = ""
    for row in rows:
        cells = await row.querySelectorAll('td')
        cell1_value = await page.evaluate('(element) => element.textContent', cells[1])
        cell2_value = await page.evaluate('(element) => element.textContent', cells[2])
        current_user = cell1_value if cell1_value else current_user
        split_text = re.split('מרכז שרון - |מרכז שרון – ', cell2_value)
        if len(split_text) > 1:
            team_list[split_text[1]].add(cell1_value if cell1_value else current_user)

    # in case it's vacation team members page, team leader is not necessarily in the list (usually not)
    if url_page == ULR_VACATION_TEAM_MEMBERS:
        await page.close()
        return team_list

    # filter out all unrelated units such as (מאגר משפחות לליווי, שם הסניף וכו׳)
    await page.close()
    return {key: value for key, value in team_list.items() if key in team_list[key]}


def apply_borders_to_all_teams(sheet, start_row, team_list):
    # Find the column index of the header
    column_index = TEAMS_SHEET_NAME_HEADER_COLUMN_INDEX # __find_header_index(sheet, header_name)

    # Iterate over each team leader in the all_teams
    for team_leader in team_list.keys():
        # Find the first and last row of the team members under the team leader
        team_leader_row, last_team_member_row = __find_first_and_last_team_member_rows(sheet, start_row, team_leader,
                                                                                       column_index)

        # Apply border to the team table
        __apply_border_to_team_table(sheet, team_leader_row, last_team_member_row, column_index, 7)


def update_wb_active_team_members(sheet, start_row, team_list):
    column_index = TEAMS_SHEET_NAME_HEADER_COLUMN_INDEX

    # Add new cells to the column
    i = start_row
    for key, values in team_list.items():
        team_lead_cell = sheet.cell(row=i, column=column_index)
        set_cell_value(team_lead_cell, key, LIGHT_BLUE_FILL)
        set_cell_value(sheet.cell(row=i, column=column_index + 1), CHECK_MARK)

        i += 1
        for value in values:
            set_cell_value(sheet.cell(row=i, column=column_index), value)
            set_cell_value(sheet.cell(row=i, column=column_index + 1), CHECK_MARK)
            i += 1

        i += 2  # Leave two empty rows to separate between team tables in the excel file


def update_wb_vacation_team_members(sheet, start_row, team_list):
    column_index = TEAMS_SHEET_NAME_HEADER_COLUMN_INDEX

    # Iterate over each team leader in the team_list
    for team_leader, team_members in team_list.items():
        # Find the first and last row of the team members under the team leader
        _, last_team_member_row = __find_first_and_last_team_member_rows(sheet, start_row, team_leader, column_index)

        # Insert the team members who are on vacation after the last row of the team members
        for i, team_member in enumerate(team_members, start=1):
            sheet.insert_rows(last_team_member_row + i)
            set_cell_value(sheet.cell(row=last_team_member_row + i, column=column_index), team_member)
            set_cell_value(sheet.cell(row=last_team_member_row + i, column=column_index + 2), CHECK_MARK)


async def collect_tutor_families(browser, unit_name, url_page, family_status):
    page = await browser.newPage()
    await page.goto(url_page)

    await filter_unit_name_with_search_button(page, unit_name, family_status)

    active_families_list = defaultdict(lambda: [])
    rows = await page.querySelectorAll('tr[id^="family_"]')

    team_leader_to_families = defaultdict(lambda: [])
    for row in rows:
        cells = await row.querySelectorAll('td')
        assigned_to = await page.evaluate('(element) => element.textContent', cells[TUTOR_COLUMN_IN_TEAMS_SHEET])
        family_name = await page.evaluate('(element) => element.textContent', cells[0])
        a_element = await cells[0].querySelector('a')
        family_link = await page.evaluate('(element) => element.href', a_element)
        unit_full_name = await page.evaluate('(element) => element.textContent', cells[1])
        team_leader = re.split('מרכז שרון - |מרכז שרון – ', unit_full_name)
        if len(team_leader) > 1:
            team_leader_to_families[team_leader[1]].append(family_name)
        families = active_families_list[assigned_to]
        active_families_list[assigned_to] = families + [(family_name, family_link)]

    await page.close()
    return active_families_list, team_leader_to_families


def update_wb_families_status(sheet, start_row, family_count_column_shift, family_list_column_shift, tutor_to_families):
    # Find the column index of the header
    column_index = TEAMS_SHEET_NAME_HEADER_COLUMN_INDEX # __find_header_index(sheet, header_name)

    for row in range(start_row, sheet.max_row + 1):
        # Check if the cell in the escort person column is valid for parsing
        if (sheet.cell(row=row, column=column_index).value and
                sheet.cell(row, column_index).fill.start_color.rgb == "00000000" and sheet.cell(row,
                                                                                                column_index).value != '-'):
            # Get the value of the escort person
            escort_person = sheet.cell(row=row, column=column_index).value
            # the escort families is a tuple of the number of families and a list of the families
            escort_families = tutor_to_families.get(escort_person, [])
            set_cell_value(sheet.cell(row=row, column=column_index + family_count_column_shift), len(escort_families))

            # Insert new rows for each additional family and create a hyperlink in each cell
            for i, (name, link) in enumerate(escort_families, start=1):
                if i > 1:
                    sheet.insert_rows(row + i - 1)
                    set_cell_value(sheet.cell(row + i - 1, column_index), '-')
                    set_cell_value(sheet.cell(row + i - 1, column_index + family_count_column_shift), '-')

                families_cell = sheet.cell(row=row + i - 1, column=column_index + family_list_column_shift)

                cell_value = f'=HYPERLINK("{link}", "{name}")'
                set_cell_value(families_cell, cell_value, font=Font(color=Color(rgb="0000FF"), underline='single'))

                families_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                # Adjust the height of the row to fit the content of the cell
                sheet.row_dimensions[
                    row + i - 1].height = 30  # Estimate the height based on the number of lines in the cell

    # Adjust the width of the ENTIRE column to be twice as wide as its current width
    column_letter = openpyxl.utils.get_column_letter(column_index + family_list_column_shift)
    sheet.column_dimensions[column_letter].width = sheet.column_dimensions[column_letter].width * 2


def insert_totals(sheet, start_row):
    # Find the column index of the header
    column_index = TEAMS_SHEET_NAME_HEADER_COLUMN_INDEX #__find_header_index(sheet, header_name)

    # Initialize counters
    total_counter = 0
    active_counter = 0
    vacation_counter = 0
    blank_rows_count = 0

    current_team_first_row = start_row
    for row in range(start_row, sheet.max_row + 1):
        if row == current_team_first_row:
            continue # skip the team leader row
        # Check if the cell in the active column and the vacation column has a check mark
        if sheet.cell(row=row, column=column_index).value is not None:
            if sheet.cell(row=row, column=column_index).value != '-':
                total_counter += 1
                if sheet.cell(row=row, column=column_index + 1).value == CHECK_MARK:
                    active_counter += 1
                if sheet.cell(row=row, column=column_index + 2).value == CHECK_MARK:
                    vacation_counter += 1
                blank_rows_count = 0
        else:  # done iterating all team members, reached  a blank row (total line or seperator between teams)
            if blank_rows_count == 0:
                set_cell_value(sheet.cell(row=row, column=column_index), total_counter, LIGHT_BLUE_FILL)
                set_cell_value(sheet.cell(row=row, column=column_index + 1), active_counter, LIGHT_BLUE_FILL)
                set_cell_value(sheet.cell(row=row, column=column_index + 2), vacation_counter, LIGHT_BLUE_FILL)
                # just color rest of the row (for cosmetic reasons)
                _color_empty_cells_in_totals_line(sheet, row, column_index)

                total_counter = 0
                active_counter = 0
                vacation_counter = 0
                _add_families_counters_totals(sheet, current_team_first_row, row, column_index)

                current_team_first_row = row + 2
            blank_rows_count += 1
            if blank_rows_count == 2:
                row += 2
            elif blank_rows_count > 2:
                _add_all_branch_totals(sheet, start_row, row, column_index)
                break

    # Insert a new row at the end
    sheet.append([])


def _color_empty_cells_in_totals_line(sheet, row, column_index):
    set_cell_value(sheet.cell(row=row, column=column_index + 4), '', LIGHT_BLUE_FILL)
    set_cell_value(sheet.cell(row=row, column=column_index + 6), '', LIGHT_BLUE_FILL)
    set_cell_value(sheet.cell(row=row, column=column_index + 7), '', LIGHT_BLUE_FILL)


def _add_families_counters_totals(sheet, start_row, end_row, column_index):
    # add the sum of all the counters (active families counter and ready_families_counter) in the proper columns
    column_letter = openpyxl.utils.get_column_letter(column_index+READY_FAMILIES_SUM_COLUMN_DIFF)
    set_cell_value(sheet.cell(row=end_row, column=column_index+READY_FAMILIES_SUM_COLUMN_DIFF),
                   f'=SUM({column_letter}{start_row}:{column_letter}{end_row-1})', LIGHT_BLUE_FILL)
    column_letter = openpyxl.utils.get_column_letter(column_index + ACTIVE_FAMILIES_SUM_COLUMN_DIFF)
    set_cell_value(sheet.cell(row=end_row, column=column_index + ACTIVE_FAMILIES_SUM_COLUMN_DIFF),
                   f'=SUM({column_letter}{start_row}:{column_letter}{end_row-1})', LIGHT_BLUE_FILL)


def _add_all_branch_totals(sheet, start_row, end_row, column_index):
    set_cell_value(sheet.cell(row=end_row, column=column_index-1), "סה״כ בסניף", LIGHT_BLUE_FILL, adjust_width=True)

    # set the escorts (active and vacation) counters
    for column in [column_index, column_index + 1, column_index + 2]:
        set_sum_formula_to_cell(sheet, start_row, end_row, column)

    # for the families counters the SUM formula counts also the teams total lines, so we need to divide it by 2
    set_sum_formula_to_cell(sheet, start_row, end_row, column_index + READY_FAMILIES_SUM_COLUMN_DIFF, divide_by_2=True)
    set_sum_formula_to_cell(sheet, start_row, end_row, column_index + ACTIVE_FAMILIES_SUM_COLUMN_DIFF, divide_by_2=True)

    # just color rest of the row (for cosmetic reasons)
    _color_empty_cells_in_totals_line(sheet, end_row, column_index)


def __find_first_and_last_team_member_rows(sheet, start_row, team_leader, column_index):
    team_leader_row = None
    for row in range(start_row, sheet.max_row + 1):
        if sheet.cell(row=row, column=column_index).value == team_leader:
            team_leader_row = row
            break

    if team_leader_row is None:
        print(f"Error team leader '{team_leader}' not found.")
        return None, None

    # Find the last row of the team members under the team leader
    last_team_member_row = team_leader_row
    while sheet.cell(row=last_team_member_row + 1, column=column_index).value is not None:
        last_team_member_row += 1

    return team_leader_row, last_team_member_row
